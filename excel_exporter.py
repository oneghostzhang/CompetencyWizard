"""
competency_wizard/excel_exporter.py  v2.0
將職能說明書資料輸出為 Excel 檔案（對齊 ICAP 職能基準書欄位格式）

Sheets:
  1. 職能說明書   — 主要職責 / 工作任務 / 工作產出 / 行為指標 / 職能等級
  2. 知識清單     — 對應知識項目
  3. 技能清單     — 對應技能項目
  4. 補充說明     — 員工補充備注
"""

from pathlib import Path
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─────────────────────────────────────────
# 顏色常數
# ─────────────────────────────────────────
C_HEADER_BG    = "2F5496"   # 深藍表頭
C_HEADER_FONT  = "FFFFFF"   # 白字
C_SUBHEAD_BG   = "D9E1F2"   # 淺藍小節
C_RESP_BG      = "E8F4FD"   # 主責列背景（淺藍）
C_TASK_BG      = "FFFFFF"   # 任務列背景（白）
C_BEHAVIOR_BG  = "E2EFDA"   # 行為指標（淺綠）
C_KNOWLEDGE_BG = "FFF2CC"   # 知識（淺黃）
C_SKILL_BG     = "F4E6FF"   # 技能（淺紫）
C_META_BG      = "F8F9FA"   # 基本資訊區（淺灰）


# ─────────────────────────────────────────
# 共用工具
# ─────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _style(cell, bg: str | None = None, bold: bool = False,
           align: str = "left", wrap: bool = False, font_size: int = 10,
           font_color: str = "000000"):
    if bg:
        cell.fill = _fill(bg)
    cell.font  = Font(name="Microsoft JhengHei", bold=bold,
                      size=font_size, color=font_color)
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin_border()


def _header_row(ws, row: int, labels: list, bg: str = C_HEADER_BG):
    for col, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=label)
        _style(c, bg=bg, bold=True, align="center", font_color=C_HEADER_FONT
               if bg == C_HEADER_BG else "1A3A5C")


def _write(ws, coord, value, **kwargs):
    ws[coord] = value
    _style(ws[coord], **kwargs)


# ─────────────────────────────────────────
# Sheet 1: 職能說明書
# ─────────────────────────────────────────

def _sheet_competency(wb, data: dict):
    ws = wb.active
    ws.title = "職能說明書"

    # 欄寬
    col_widths = [10, 14, 10, 28, 36, 40, 10]
    headers    = ["主責代碼", "主責名稱", "任務代碼", "任務名稱", "工作產出", "行為指標", "職能等級"]
    for i, (w, h) in enumerate(zip(col_widths, headers), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 基本資訊區 ──────────────────────────────────────────────────────────
    info_pairs = [
        ("職業名稱",   data.get("position", "")),
        ("職能等級",   str(data.get("level", ""))),
        ("職能基準",   data.get("standard_name", "（未使用基準）")),
        ("基準代碼",   data.get("standard_code", "")),
        ("建立日期",   datetime.now().strftime("%Y-%m-%d")),
    ]
    row = 1
    ws.merge_cells(f"A{row}:G{row}")
    _write(ws, f"A{row}", "職能說明書",
           bg=C_HEADER_BG, bold=True, align="center",
           font_color=C_HEADER_FONT, font_size=13)
    ws.row_dimensions[row].height = 28
    row += 1

    for label, value in info_pairs:
        ws[f"A{row}"] = label
        _style(ws[f"A{row}"], bg=C_META_BG, bold=True, align="right")
        ws.merge_cells(f"B{row}:G{row}")
        _write(ws, f"B{row}", value, bg=C_META_BG)
        row += 1

    row += 1  # 空一列

    # ── 欄位標題 ──────────────────────────────────────────────────────────
    _header_row(ws, row, headers)
    ws.row_dimensions[row].height = 22
    row += 1

    # ── 資料列 ────────────────────────────────────────────────────────────
    rows_data = data.get("rows", [])
    for r in rows_data:
        behaviors = r.get("behavior_accepted") or []
        behavior_str = "\n".join(f"・{b}" for b in behaviors) if behaviors else ""

        # 一行一個任務（行為指標合併在同格，換行顯示）
        values = [
            r.get("resp_code", ""),
            r.get("resp_name", ""),
            r.get("task_code", ""),
            r.get("task_name", ""),
            r.get("output", ""),
            behavior_str,
            str(r.get("level", "")),
        ]
        bgs = [C_RESP_BG, C_RESP_BG, C_TASK_BG, C_TASK_BG,
               C_TASK_BG, C_BEHAVIOR_BG, C_TASK_BG]
        aligns = ["center", "left", "center", "left", "left", "left", "center"]

        for col, (val, bg, aln) in enumerate(zip(values, bgs, aligns), 1):
            c = ws.cell(row=row, column=col, value=val)
            _style(c, bg=bg, align=aln, wrap=True)

        if behavior_str:
            line_count = max(1, behavior_str.count("\n") + 1)
            ws.row_dimensions[row].height = max(18, line_count * 16)
        row += 1

    # 凍結首列（標題）
    ws.freeze_panes = ws.cell(row=9, column=1)


# ─────────────────────────────────────────
# Sheet 2: 知識清單
# ─────────────────────────────────────────

def _collect_ks(rows: list, field: str):
    """彙整知識或技能：{key: {code, name, tasks:[...]}}，保留插入順序。"""
    from collections import OrderedDict
    items: dict = OrderedDict()
    for r in rows:
        task_code = r.get("task_code", "")
        for entry in r.get(field, []):
            if isinstance(entry, dict):
                code = entry.get("code", "")
                name = entry.get("name", "")
            elif isinstance(entry, str):
                code = entry; name = entry
            else:
                continue
            key = code or name
            if not key:
                continue
            if key not in items:
                items[key] = {"code": code, "name": name, "tasks": []}
            if task_code and task_code not in items[key]["tasks"]:
                items[key]["tasks"].append(task_code)
    return list(items.values())


def _sheet_knowledge(wb, data: dict):
    ws = wb.create_sheet("知識清單")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28

    _header_row(ws, 1, ["知識代碼", "知識名稱", "對應任務"])
    row = 2

    items = _collect_ks(data.get("rows", []), "_knowledge")
    for item in items:
        ws.cell(row=row, column=1, value=item["code"])
        _style(ws.cell(row=row, column=1), bg=C_KNOWLEDGE_BG, align="center")
        ws.cell(row=row, column=2, value=item["name"])
        _style(ws.cell(row=row, column=2), bg=C_KNOWLEDGE_BG)
        ws.cell(row=row, column=3, value="、".join(item["tasks"]))
        _style(ws.cell(row=row, column=3), bg=C_KNOWLEDGE_BG, align="center")
        row += 1

    if row == 2:
        ws.merge_cells("A2:C2")
        _write(ws, "A2", "（未填寫知識項目）", bg=C_META_BG, align="center")


# ─────────────────────────────────────────
# Sheet 3: 技能清單
# ─────────────────────────────────────────

def _sheet_skills(wb, data: dict):
    ws = wb.create_sheet("技能清單")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28

    _header_row(ws, 1, ["技能代碼", "技能名稱", "對應任務"])
    row = 2

    items = _collect_ks(data.get("rows", []), "_skills")
    for item in items:
        ws.cell(row=row, column=1, value=item["code"])
        _style(ws.cell(row=row, column=1), bg=C_SKILL_BG, align="center")
        ws.cell(row=row, column=2, value=item["name"])
        _style(ws.cell(row=row, column=2), bg=C_SKILL_BG)
        ws.cell(row=row, column=3, value="、".join(item["tasks"]))
        _style(ws.cell(row=row, column=3), bg=C_SKILL_BG, align="center")
        row += 1

    if row == 2:
        ws.merge_cells("A2:C2")
        _write(ws, "A2", "（未填寫技能項目）", bg=C_META_BG, align="center")


# ─────────────────────────────────────────
# ─────────────────────────────────────────
# Sheet 4: 態度清單
# ─────────────────────────────────────────

C_ATTITUDE_BG = "FFE4E1"   # 淺粉紅（態度）

def _sheet_attitudes(wb, data: dict):
    ws = wb.create_sheet("態度清單")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 60

    _header_row(ws, 1, ["態度代碼", "態度名稱", "說明"])
    row = 2

    for a in data.get("attitudes", []):
        if not isinstance(a, dict):
            continue
        code = a.get("code", "")
        name = a.get("name", "")
        desc = a.get("description", "")
        if not (code or name):
            continue
        ws.cell(row=row, column=1, value=code)
        _style(ws.cell(row=row, column=1), bg=C_ATTITUDE_BG, align="center")
        ws.cell(row=row, column=2, value=name)
        _style(ws.cell(row=row, column=2), bg=C_ATTITUDE_BG, bold=True)
        ws.cell(row=row, column=3, value=desc)
        _style(ws.cell(row=row, column=3), bg=C_ATTITUDE_BG, wrap=True)
        ws.row_dimensions[row].height = max(18, len(desc) // 30 * 16)
        row += 1

    if row == 2:
        ws.merge_cells("A2:C2")
        _write(ws, "A2", "（無態度職能內涵資料）", bg=C_META_BG, align="center")


# ─────────────────────────────────────────
# Sheet 5: 補充說明
# ─────────────────────────────────────────

def _sheet_supplement(wb, data: dict):
    ws = wb.create_sheet("補充說明")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60

    ws.merge_cells("A1:B1")
    _write(ws, "A1", "補充說明", bg=C_HEADER_BG, bold=True,
           align="center", font_color=C_HEADER_FONT, font_size=12)

    pairs = [
        ("職業名稱",    data.get("position", "")),
        ("職能等級",    str(data.get("level", ""))),
        ("職能基準名稱", data.get("standard_name", "")),
        ("建立日期",    datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("補充說明",    data.get("supplement", "")),
    ]
    for row_idx, (label, value) in enumerate(pairs, 2):
        ws[f"A{row_idx}"] = label
        _style(ws[f"A{row_idx}"], bg=C_META_BG, bold=True, align="right")
        ws[f"B{row_idx}"] = value
        _style(ws[f"B{row_idx}"], bg="FFFFFF", wrap=True)
        ws.row_dimensions[row_idx].height = 20

    # 工作任務摘要
    row = len(pairs) + 3
    ws.merge_cells(f"A{row}:B{row}")
    _write(ws, f"A{row}", "工作任務摘要", bg=C_SUBHEAD_BG, bold=True)
    row += 1

    for r in data.get("rows", []):
        behaviors = r.get("behavior_accepted") or []
        label_txt = f"{r.get('task_code','')}  {r.get('task_name','')}"
        behaviors_txt = "；".join(behaviors) if behaviors else "（未生成行為指標）"
        ws[f"A{row}"] = label_txt
        _style(ws[f"A{row}"], bg=C_TASK_BG, bold=True)
        ws[f"B{row}"] = behaviors_txt
        _style(ws[f"B{row}"], bg=C_TASK_BG, wrap=True)
        ws.row_dimensions[row].height = max(18, len(behaviors) * 16)
        row += 1


# ─────────────────────────────────────────
# 公開進入點
# ─────────────────────────────────────────

def export_competency(
    data: dict,
    output_path: Optional[Path] = None,
) -> Path:
    """
    輸出職能說明書 Excel（v2.0 格式）。

    Args:
        data: {
            position, level, standard_code, standard_name, supplement,
            rows: [{resp_code, resp_name, task_code, task_name, output,
                    level, _knowledge, _skills, behavior_accepted}]
        }
        output_path: 輸出路徑（None 則自動命名）

    Returns:
        實際輸出的 Path
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("缺少 openpyxl，請執行 pip install openpyxl")

    if output_path is None:
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe = data.get("position", "職能說明書").replace("/", "_")
        output_path = Path.cwd() / f"{safe}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    _sheet_competency(wb, data)
    _sheet_knowledge(wb, data)
    _sheet_skills(wb, data)
    _sheet_attitudes(wb, data)
    _sheet_supplement(wb, data)

    wb.save(str(output_path))
    return output_path
